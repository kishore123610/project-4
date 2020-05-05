# -*- coding: utf-8 -*-
"""
Created on Sun Oct  7 11:54:41 2018

@author: Cherukuri Amul
"""
#
#f=open("F:/python/file.txt")
##print(f.readline())
##print(f.read())
##print(f.readlines())
##print(f.readlines(10))
#f.close()
#=========
#f=open("F:/python/file.txt")
#for i in f:
#    print(i)
#==============



#write mode
#f=open("F:/python/files1.txt",'w')
#f.write('amul \n')
#f.write('hero')
#f.close()
#
#
#
#
##write in excel files using csv
#f=open("F:/python/files1.csv",'w')
#f.write('amul \n')
#f.write('hero')







#append
#
#f=open("F:/python/files1.csv",'a')
#f.write('good boy')
#f.close()




#append mode using iteration
#list=['\n amul,\t','amul@gmail.com \n','jasu,\t','jasu@gmail.com \n']
#f=open("F:/python/files1.csv",'a')
#for i in list:
#    f.write(i)
#f.close()
    

#without file handling
#with open("F:/python/files1.csv") as f:
#    print(f.read())

#
#with open("F:/python/files1.csv",'w') as f:
#    print(f.write('amul'))



#with open("F:/python/files1.csv") as f:
#    a=f.read()
#    print(a)
#
#with open("F:/python/files12.csv",'w') as f:
#    print(f.write(a))
#    
#    
#copying file without shutil
#with open("F:/python/files1.csv") as f:
#    a=f.read()
#    print(a)
#    with open("F:/python/files123.csv",'w') as f:
#        print(f.write(a))



#######################################################################################
#inplace of the above method we can use shutil to copy 
#import shutil
#shutil.copy('F:/python/files1.csv','F:/python/files1234.csv')


#deleting file
#import os
#os.remove("F:/python/files123.csv")

##r+(files wont be overwriten)
#f=open("F:/python/files12.csv",'r+')
#print(f.read())
#f.write("hi")
#f.write('hello')
#f.close()


#w+(files will be overwritten)
#f=open("F:/python/files12.csv",'w+')
#
#f.write("hi")
#f.write('hello')
#print(f.read())
#f.close()


#import pickle
#f=open("F:/python/filebinary.txt",'wb')
#pickle.dump(11,f)
#pickle.dump("this is a line",f)
#pickle.dump([1,2,3,4],f)
#f.close()

#print("+++++++++++++++++++++++++")
#
#import pickle
#f=open("F:/python/filebinary.txt",'rb')
#for i in f:
#    print(pickle.load(i))
##print(pickle.load(f))
##print(pickle.load(f))
##"this is a line"
##print(pickle.load(f))
#f.close()


#write in excel files using csv
#f=open("F:/python/files1.csv",'wb')
#f.write('amul \n')
#f.write('hero')






#using input method writing into data


#a=input("Insert an element")
#f=open("F:/python/files1234.csv",'w')
#f.write(a)
#f.close()





#from openpyxl import Workbook
#wb = Workbook()
#ws = wb.active
#ws['A1'] = 42
#
#
#wb.save("‪info.xlsx")


from openpyxl import load_workbook
wb = load_workbook(filename = '‪info.xlsx')

print(wb)


